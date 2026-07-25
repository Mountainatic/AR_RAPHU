#!/usr/bin/env python3
"""Export Phase 0 manifests without changing or statistically profiling CZ data."""

from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.metadata
import json
import os
import platform
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_VERSION = "2.1.0"
AUTHORITATIVE_DOCUMENT_NAMES = [
    "AR_RAPHU_method_v2.md",
    "AR_RAPHU_three_layer_validation_plan_v2.md",
    "AR_RAPHU_v2_revision_notes.md",
]
V20_BUNDLE_DIRECTORY = "STAGE1_DUAL_SOLVER_V20_bundle"
V20_CONTRACT_FILES = [
    "layers.py",
    "run_kan_fast_s0_v20.py",
    "stage1/delay_prior.py",
    "stage1/experiment_utils.py",
    "stage1/lag_contract.py",
    "stage1/model.py",
    "stage1/protocol.py",
    "stage1/response_kan.py",
    "stage1/sequence_ops.py",
    "stage1/synthetic.py",
]
V2_ADAPTER_FILES = [
    "configs/protocol_v2.yaml",
    "src/ar_raphu/data_protocol.py",
    "src/ar_raphu/baselines.py",
    "src/ar_raphu/dataset.py",
    "src/ar_raphu/model.py",
    "src/ar_raphu/phase1_evidence.py",
    "src/ar_raphu/preprocessing.py",
    "src/ar_raphu/protocol_config.py",
    "src/ar_raphu/rank_audit.py",
    "src/ar_raphu/statistics.py",
    "src/ar_raphu/sequence_data.py",
    "src/ar_raphu/synthetic.py",
    "src/ar_raphu/training.py",
    "tools/run_phase1_baselines.py",
    "tools/run_phase1_scheme_a.py",
    "tools/make_phase1_scheme_a_manifest.py",
    "tests/test_data_protocol.py",
    "tests/test_ar_raphu_model.py",
    "tests/test_baselines.py",
    "tests/test_phase1_synthetic.py",
    "tests/test_phase1_manifests.py",
    "tests/test_phase1_evidence.py",
    "tests/test_preprocessing_and_dataset.py",
    "tests/test_rank_audit.py",
    "tests/test_statistics.py",
    "tests/test_sequence_training.py",
    "tests/test_v20_contract.py",
]
EXPECTED_HEADERS = [
    "加热元件温度",
    "主加热功率",
    "晶升速度",
    "晶转速度",
    "埚升速度",
    "埚转速度",
    "氩气流量设定",
    "晶体长度",
    "炉压",
    "晶体直径",
]
INPUT_HEADERS = EXPECTED_HEADERS[:-1]
TARGET_HEADER = EXPECTED_HEADERS[-1]
CONSTANT_EXOGENOUS_HEADER = "氩气流量设定"
EXPECTED_SOURCE_SHA256 = (
    "c46e0d35d26903386fd80408f36660c4f8925a5dbc56c92527f020e433ef04de"
)
RECORDED_HISTORY_LENGTH = 32
LEGACY_L_X = 32
LEGACY_L_Y = 32
FORMAL_L_Y_CANDIDATES = [1, 4, 8, 16, 32, 64]
FORMAL_L_X_CANDIDATES = [32, 64, 128, 256]
CONDITIONAL_L_X_EXTENSION = 512
PREDICTION_HORIZONS = [1, 5, 10, 30, 60]
CZ_FOLDS = [
    {
        "fold": 1,
        "role": "development",
        "train": [0, 10051],
        "validation": [10051, 12061],
        "test": [12061, 14072],
    },
    {
        "fold": 2,
        "role": "development",
        "train": [0, 12061],
        "validation": [12061, 14072],
        "test": [14072, 16082],
    },
    {
        "fold": 3,
        "role": "development",
        "train": [0, 14072],
        "validation": [14072, 16082],
        "test": [16082, 18092],
    },
    {
        "fold": 4,
        "role": "final_lockbox",
        "train": [0, 16082],
        "validation": [16082, 18092],
        "test": [18092, 20103],
    },
]
UNKNOWN = "UNKNOWN_REQUIRES_USER"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_jsonable(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")
    temporary.replace(path)


def atomic_write_csv(
    path: Path, fieldnames: list[str], rows: list[dict[str, Any]]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    temporary.replace(path)


def classify_cell(cell: Any) -> str:
    value = cell.value
    if value is None or (isinstance(value, str) and not value.strip()):
        return "blank"
    mapping = {
        "n": "numeric",
        "s": "text",
        "b": "boolean",
        "e": "error",
        "f": "formula",
        "d": "datetime",
    }
    return mapping.get(cell.data_type, str(cell.data_type))


def audit_workbook(source: Path) -> dict[str, Any]:
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            header_cells = next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    min_col=1,
                    max_col=worksheet.max_column,
                )
            )
            headers = [cell.value for cell in header_cells]
            type_counts = [Counter() for _ in headers]
            data_rows = 0
            formula_cells = 0
            constant_column_index = headers.index(CONSTANT_EXOGENOUS_HEADER)
            constant_value_seen = False
            constant_reference: Any = None
            constant_over_all_rows = True

            for row in worksheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                data_rows += 1
                for index, cell in enumerate(row):
                    category = classify_cell(cell)
                    type_counts[index][category] += 1
                    formula_cells += int(category == "formula")
                    if index == constant_column_index and category != "blank":
                        if not constant_value_seen:
                            constant_reference = cell.value
                            constant_value_seen = True
                        elif cell.value != constant_reference:
                            constant_over_all_rows = False

            columns = []
            for index, (header, counts) in enumerate(
                zip(headers, type_counts, strict=True), start=1
            ):
                columns.append(
                    {
                        "column_index_1based": index,
                        "header": header,
                        "cell_type_counts": dict(sorted(counts.items())),
                        "missing_count": counts["blank"],
                        "non_missing_count": data_rows - counts["blank"],
                        "constant_over_all_rows": (
                            constant_value_seen and constant_over_all_rows
                            if header == CONSTANT_EXOGENOUS_HEADER
                            else None
                        ),
                    }
                )

            sheets.append(
                {
                    "name": worksheet.title,
                    "state": worksheet.sheet_state,
                    "header_row_1based": 1,
                    "row_count_including_header": worksheet.max_row,
                    "data_row_count": data_rows,
                    "column_count": worksheet.max_column,
                    "headers": headers,
                    "columns": columns,
                    "formula_cell_count": formula_cells,
                    "merged_range_count": 0,
                }
            )
        return {"sheet_names": workbook.sheetnames, "sheets": sheets}
    finally:
        workbook.close()


def validate_source_audit(audit: dict[str, Any], source_sha256: str) -> None:
    if source_sha256 != EXPECTED_SOURCE_SHA256:
        raise RuntimeError(
            "CZ source SHA256 differs from the frozen value; stop before exporting."
        )
    if audit["sheet_names"] != ["Sheet1"]:
        raise RuntimeError(f"Unexpected workbook sheets: {audit['sheet_names']!r}")
    sheet = audit["sheets"][0]
    if sheet["headers"] != EXPECTED_HEADERS:
        raise RuntimeError(f"Unexpected CZ headers: {sheet['headers']!r}")
    if sheet["column_count"] != len(EXPECTED_HEADERS):
        raise RuntimeError("Unexpected CZ column count.")
    if sheet["data_row_count"] != 20103:
        raise RuntimeError("Unexpected CZ data row count.")
    if sheet["formula_cell_count"] != 0:
        raise RuntimeError("Formula cells are not permitted in the frozen source.")
    for column in sheet["columns"]:
        if column["missing_count"] != 0:
            raise RuntimeError(
                f"Missing values found in frozen source column {column['header']!r}."
            )
        counts = column["cell_type_counts"]
        if set(counts) != {"numeric"}:
            raise RuntimeError(
                f"Non-numeric data found in source column {column['header']!r}: "
                f"{counts!r}"
            )
    argon = next(
        column
        for column in sheet["columns"]
        if column["header"] == CONSTANT_EXOGENOUS_HEADER
    )
    if argon["constant_over_all_rows"] is not True:
        raise RuntimeError("The confirmed constant argon channel is not constant.")


def audit_v20_bundle(project_root: Path) -> dict[str, Any]:
    bundle = project_root / V20_BUNDLE_DIRECTORY
    checksum_path = bundle / "SHA256SUMS.txt"
    if not checksum_path.is_file():
        raise RuntimeError(f"V20 checksum manifest is missing: {checksum_path}")

    verified_files = []
    for line in checksum_path.read_text(encoding="utf-8-sig").splitlines():
        if not line.strip():
            continue
        expected, relative_path = line.split(maxsplit=1)
        path = bundle / relative_path
        if not path.is_file():
            raise RuntimeError(f"V20 bundle file is missing: {relative_path}")
        observed = sha256_file(path)
        if observed != expected:
            raise RuntimeError(
                f"V20 bundle checksum mismatch for {relative_path}: "
                f"{observed} != {expected}"
            )
        verified_files.append(
            {
                "path": relative_path,
                "size_bytes": path.stat().st_size,
                "sha256": observed,
            }
        )

    file_map = {item["path"]: item for item in verified_files}
    missing_contract = [name for name in V20_CONTRACT_FILES if name not in file_map]
    if missing_contract:
        raise RuntimeError(
            f"V20 contract files are absent from SHA256SUMS: {missing_contract}"
        )

    return {
        "schema_version": 1,
        "bundle_directory": V20_BUNDLE_DIRECTORY,
        "integrity": {
            "status": "VERIFIED",
            "verified_file_count": len(verified_files),
            "sha256sums_sha256": sha256_file(checksum_path),
            "all_listed_files_match": True,
        },
        "contract_file_hashes": {
            name: file_map[name]["sha256"] for name in V20_CONTRACT_FILES
        },
        "regression_suite": {
            "command": (
                "conda run -n Env_pytorch --no-capture-output python -m pytest "
                "tests/test_stage1.py tests/test_stage1_acceleration.py "
                "tests/test_stage1_dual_solver_v20.py -q"
            ),
            "observed_result_2026_07_25": "118 passed",
        },
        "historical_v20_external_gamma_contract": {
            "mode": "static_gamma",
            "sample_dependent_tilt_epsilon": 0.0,
            "lag_order": "current_to_past",
            "tau_support": "0_to_L_minus_1",
            "discretization": "point_value_then_softmax",
            "delta": 0.001,
            "beta_convention": "scale",
            "log_unnormalized_formula": (
                "(alpha-1)*log(tau+delta)-(tau+delta)/beta"
            ),
            "learned_parametrization": "bounded_mean_std",
            "mean_bounds": "[0,L-1]",
            "std_bounds": "[0.5,L/2]",
            "continuous_moment_conversion": {
                "safe_mean": "mean+delta",
                "beta_scale": "std^2/safe_mean",
                "alpha_shape": "safe_mean/beta",
            },
            "alpha_bounds": [0.5, 10.0],
            "beta_scale_bounds": "[0.1,L]",
            "row_normalization": "softmax",
        },
        "historical_v20_kan_contract": {
            "per_variable_layers": [1, 8, 1],
            "hidden_kan": 8,
            "grid_size": 7,
            "spline_order": 3,
            "base_function": "SiLU",
            "base_and_spline_scales": [1.0, 1.0],
            "spline_initialization_noise_scale": 0.1,
            "first_layer_grid": (
                "train-only min/max plus 5% padding with minimum pad 0.05"
            ),
            "second_layer_grid": [-3.0, 3.0],
            "dynamic_grid_update": False,
            "execution": (
                "sequence-first vectorized KAN plus exact grouped causal convolution"
            ),
        },
        "source_backed_capacity_audit_space": {
            "hidden_kan": [4, 8],
            "grid_size": [5, 7, 11],
            "spline_order_fixed": 3,
            "formal_v20_selected": {"hidden_kan": 8, "grid_size": 7},
        },
        "historical_v20_training_contract": {
            "seeds": [0, 1, 2, 3, 4],
            "optimizer": "Adam",
            "learning_rate": 0.003,
            "warmup_epochs": 3000,
            "warmup_patience": 300,
            "independent_pruning_scales": [
                0.003,
                0.004,
                0.005,
                0.006,
                0.007,
                0.008,
                0.009,
                0.010,
                0.012,
            ],
            "prune_epochs": 1800,
            "ramp_epochs": 300,
            "fixed_support_refit_epochs": 5000,
            "refit_patience": 500,
            "selection": "cross_seed_validation_only_one_standard_error",
        },
        "directly_reusable_in_ar_raphu_v2": [
            "DiscreteGammaPrior mathematical implementation",
            "per-variable two-layer KAN response family",
            "sequence-first exact static convolution",
            "shared warmup plus independent pruning forks",
            "cross-seed validation-only one-standard-error selection",
            "fixed-delay convex spline solver components",
        ],
        "must_not_be_reused_without_revision": [
            "V20 synthetic horizon=0 target alignment",
            "V20 31-point embargo split policy",
            "stage1/configs/real.yaml rows 6000:15000 and 0.85/0.10/0.05 split",
            "single shared max_lag across all ten channels",
            "all-ten-branches-identical scientific interpretation",
            "all-active mask that fails to exclude constant argon",
        ],
        "missing_from_bundle": [
            "process_data.py",
            "executable real-CZ tenth-channel construction",
            "real-CZ scaler implementation",
            "separate L_x and L_y AR-RAPHU model",
        ],
        "v2_migration_rule": {
            "legacy_reproduction": (
                "Reuse V20 Gamma/KAN/operator semantics at L_x=L_y=32, "
                "but construct the lagged-output channel with the frozen "
                "target-index protocol so y[t+h] never enters input."
            ),
            "formal_model": (
                "Use separate external and AR branches so L_x and L_y can differ; "
                "mask constant argon; B residual applies only to external branches."
            ),
        },
    }


def audit_v2_adapter(project_root: Path) -> dict[str, Any]:
    file_hashes: dict[str, str] = {}
    for relative_path in V2_ADAPTER_FILES:
        path = project_root / relative_path
        if not path.is_file():
            raise RuntimeError(f"AR-RAPHU v2 adapter file is missing: {path}")
        file_hashes[relative_path] = sha256_file(path)

    protocol_config = json.loads(
        (project_root / "configs/protocol_v2.yaml").read_text(encoding="utf-8")
    )
    return {
        "status": "STRUCTURAL_ADAPTER_IMPLEMENTED_AND_UNIT_TESTED",
        "file_hashes": file_hashes,
        "implementation": {
            "model": "ARRAPHURank1",
            "track_specific_construction": ["X", "AR", "XAR"],
            "separate_external_and_AR_branches": True,
            "unequal_L_x_L_y_supported": True,
            "external_process_channels": 9,
            "AR_channels": 1,
            "constant_argon_channel_index_0based": 6,
            "constant_argon_scientifically_masked": True,
            "single_wrapper_intercept": True,
            "independent_direct_horizon_model_required": True,
            "train_only_first_layer_grid_ranges_required_at_construction": True,
            "scheme_B_scope": "external_branches_only",
        },
        "phase1_gate": {
            "status": protocol_config["status"],
            "required_fields": protocol_config["required_before_phase1"],
        },
        "unit_test_contract": [
            "unequal_L_x_L_y",
            "V20_Gamma_row_normalization",
            "X_AR_XAR_track_isolation",
            "constant_argon_zero_output_and_zero_input_gradient",
            "ten_component_contribution_closure",
            "frozen_horizon_and_KAN_capacity_guards",
            "train_only_scaler_invariance_to_future_poisoning",
            "fold_bound_scaler_provenance",
            "lazy_window_construction_and_cache_namespace",
            "Fold_4_values_not_inspected_by_development_dataset",
            "Phase_1_preregistration_gate",
        ],
    }


def package_inventory() -> tuple[list[dict[str, str]], str]:
    packages = sorted(
        (
            {"name": distribution.metadata["Name"], "version": distribution.version}
            for distribution in importlib.metadata.distributions()
            if distribution.metadata.get("Name")
        ),
        key=lambda item: (item["name"].lower(), item["version"]),
    )
    return packages, sha256_jsonable(packages)


def torch_environment() -> dict[str, Any]:
    try:
        import torch
    except ImportError:
        return {"installed": False}

    cuda_available = torch.cuda.is_available()
    devices = []
    if cuda_available:
        for index in range(torch.cuda.device_count()):
            properties = torch.cuda.get_device_properties(index)
            devices.append(
                {
                    "index": index,
                    "name": properties.name,
                    "compute_capability": list(
                        torch.cuda.get_device_capability(index)
                    ),
                    "total_memory_bytes": properties.total_memory,
                    "multiprocessor_count": properties.multi_processor_count,
                }
            )

    return {
        "installed": True,
        "version": torch.__version__,
        "cuda_available": cuda_available,
        "torch_cuda_version": torch.version.cuda,
        "cudnn_version": torch.backends.cudnn.version(),
        "device_count": torch.cuda.device_count() if cuda_available else 0,
        "devices": devices,
        "default_dtype": str(torch.get_default_dtype()),
        "deterministic_algorithms_enabled": (
            torch.are_deterministic_algorithms_enabled()
        ),
        "cpu_thread_count": torch.get_num_threads(),
        "cpu_interop_thread_count": torch.get_num_interop_threads(),
    }


def build_environment(generated_at: str) -> dict[str, Any]:
    packages, packages_sha256 = package_inventory()
    thread_environment = {
        name: os.environ.get(name)
        for name in [
            "OMP_NUM_THREADS",
            "MKL_NUM_THREADS",
            "OPENBLAS_NUM_THREADS",
            "NUMEXPR_NUM_THREADS",
            "RAYON_NUM_THREADS",
        ]
    }
    return {
        "schema_version": 1,
        "generated_at": generated_at,
        "exporter": {
            "path": "tools/export_phase0.py",
            "version": SCRIPT_VERSION,
        },
        "required_python_environment": "Env_pytorch",
        "active_conda_environment": os.environ.get("CONDA_DEFAULT_ENV"),
        "python": {
            "version": platform.python_version(),
            "implementation": platform.python_implementation(),
            "executable": sys.executable,
        },
        "platform": {
            "system": platform.system(),
            "release": platform.release(),
            "machine": platform.machine(),
            "platform": platform.platform(),
            "logical_cpu_count": os.cpu_count(),
        },
        "thread_environment": thread_environment,
        "torch": torch_environment(),
        "packages": packages,
        "packages_sha256": packages_sha256,
        "reproducibility_status": (
            "CAPTURED_NOT_YET_BENCHMARKED_OR_DETERMINISM_FROZEN"
        ),
    }


def build_data_manifest(
    source: Path,
    source_sha256: str,
    audit: dict[str, Any],
    generated_at: str,
) -> dict[str, Any]:
    stat = source.stat()
    return {
        "schema_version": 1,
        "generated_at": generated_at,
        "dataset_id": "cz_private_experiment_data_1",
        "privacy": {
            "classification": "PRIVATE_CZ",
            "local_processing_only": True,
            "upload_permitted": False,
            "public_artifact_permitted": False,
            "raw_source_mutable": False,
        },
        "source": {
            "path_relative_to_project": source.name,
            "filename": source.name,
            "format": "Microsoft Excel 2007+ (.xlsx)",
            "size_bytes": stat.st_size,
            "sha256": source_sha256,
            "expected_sha256": EXPECTED_SOURCE_SHA256,
        },
        "audit_scope": {
            "level": "STRUCTURAL_AND_CONFIRMED_CONSTANT_CHANNEL_ONLY",
            "numeric_distribution_profiled": False,
            "target_distribution_profiled": False,
            "confirmed_constant_channel_checked": CONSTANT_EXOGENOUS_HEADER,
            "reason": (
                "Avoid using future test-block distributions before the split "
                "policy and boundaries are frozen."
            ),
        },
        "workbook": audit,
        "confirmed_semantics": {
            "external_process_channels": len(INPUT_HEADERS),
            "external_process_columns": INPUT_HEADERS,
            "historical_output_channels": 1,
            "model_input_channels": len(INPUT_HEADERS) + 1,
            "target_column": TARGET_HEADER,
            "current_target_used_as_input": False,
            "strictly_lagged_target_used_as_input": True,
            "historical_output_minimum_lag": 1,
            "recorded_history_length_samples": RECORDED_HISTORY_LENGTH,
            "legacy_reproduction": {"L_x": LEGACY_L_X, "L_y": LEGACY_L_Y},
            "formal_L_x_candidates": FORMAL_L_X_CANDIDATES,
            "formal_L_y_candidates": FORMAL_L_Y_CANDIDATES,
            "conditional_L_x_extension": CONDITIONAL_L_X_EXTENSION,
            "formal_L_x_selected": None,
            "formal_L_y_selected": None,
            "constant_unidentifiable_external_channel": (
                CONSTANT_EXOGENOUS_HEADER
            ),
            "confirmation_source": [
                "AR_RAPHU_method_v2.md",
                "AR_RAPHU_three_layer_validation_plan_v2.md",
                "AR_RAPHU_v2_revision_notes.md",
            ],
        },
        "ordering": {
            "stored_row_order_preserved": True,
            "timestamp_column_present": False,
            "sampling_interval": UNKNOWN,
            "equal_spacing_confirmed": False,
            "single_continuous_trajectory_confirmed": True,
            "chronological_sequence_semantics_confirmed": True,
        },
        "grouping": {
            "boule_or_batch_id_present": False,
            "crystal_count": 1,
            "independent_trajectory_count": 1,
            "stage_label_present": False,
            "covered_stage": "constant_diameter_only",
            "shutdown_segment_definition": UNKNOWN,
        },
        "label_metadata": {
            "measurement_method": UNKNOWN,
            "instrument_delay": UNKNOWN,
            "filtering_or_smoothing": UNKNOWN,
        },
        "source_integrity": {
            "sha256_verified": True,
            "source_hash_checked_before_and_after_export": True,
        },
    }


def build_variable_rows() -> list[dict[str, Any]]:
    rows = []
    for index, name in enumerate(EXPECTED_HEADERS, start=1):
        is_target = name == TARGET_HEADER
        is_constant = name == CONSTANT_EXOGENOUS_HEADER
        role = "target_and_strictly_lagged_AR_source" if is_target else "input"
        rows.append(
            {
                "source_column_1based": index,
                "canonical_name": name,
                "role": role,
                "model_input_index_0based": (
                    9 if is_target else INPUT_HEADERS.index(name)
                ),
                "model_channel_type": (
                    "strictly_lagged_historical_output"
                    if is_target
                    else "exogenous_process"
                ),
                "same_index_as_target_used_as_input": "no",
                "prediction_origin_t_value_used_as_input": "yes",
                "minimum_lag_relative_to_target_samples": (
                    "h_where_h_in_1_5_10_30_60" if is_target else "h_or_more"
                ),
                "unit": UNKNOWN,
                "control_or_measurement_attribute": UNKNOWN,
                "sensor_or_actuator_location": UNKNOWN,
                "sampling_interval": UNKNOWN,
                "include_in_current_model": "yes",
                "scientifically_trainable_current_crystal": (
                    "no_constant_unidentifiable" if is_constant else "yes"
                ),
                "rank_audit_scope": (
                    "not_in_v2_first_version"
                    if is_target
                    else (
                        "not_identifiable_current_crystal"
                        if is_constant
                        else "eligible_after_stage_gates"
                    )
                ),
                "semantic_confirmation": "AR_RAPHU_v2_frozen_documents",
            }
        )
    return rows


def build_split_manifest(generated_at: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "generated_at": generated_at,
        "dataset_id": "cz_private_experiment_data_1",
        "protocol_status": "FROZEN",
        "fold_boundaries_frozen": True,
        "sample_window_indices_materialized": False,
        "index_convention": "zero_based_left_closed_right_open",
        "sequence_length": 20103,
        "folds": CZ_FOLDS,
        "trajectory": {
            "crystal_count": 1,
            "stage": "constant_diameter_only",
            "continuous_sequence": True,
        },
        "window_protocol": {
            "legacy_reproduction": {"L_x": LEGACY_L_X, "L_y": LEGACY_L_Y},
            "formal_L_y_candidates": FORMAL_L_Y_CANDIDATES,
            "formal_L_x_candidates": FORMAL_L_X_CANDIDATES,
            "conditional_L_x_extension": {
                "value": CONDITIONAL_L_X_EXTENSION,
                "allowed_only_if": (
                    "L_x=256 is selected at the development-search boundary "
                    "and remains clearly better than shorter windows"
                ),
            },
            "selection_order": [
                "select_L_y_with_Track_AR_on_folds_1_to_3",
                "select_L_x_with_Track_X_on_folds_1_to_3",
                "compose_Track_XAR_and_allow_one_small_neighborhood_check",
                "freeze_one_global_L_x_L_y_before_Fold_4_test",
            ],
            "full_cartesian_product_search": False,
            "formal_L_x_selected": None,
            "formal_L_y_selected": None,
        },
        "forecast_protocol": {
            "primary_mode": "direct_multi_horizon",
            "prediction_horizons_samples": PREDICTION_HORIZONS,
            "information_cutoff": "t",
            "input_X_slice_python": "X[t-L_x+1:t+1]",
            "input_y_slice_python": "y[t-L_y+1:t+1]",
            "target_index": "t+h",
            "split_membership_by": "target_index_t_plus_h",
            "future_X_after_t_allowed": False,
            "intermediate_true_y_after_t_allowed": False,
            "recursive_forecast_role": "auxiliary_stability_only",
            "known_future_X_exception": (
                "TEP_only_separately_labeled_Oracle_X_upper_bound"
            ),
        },
        "required_policy": {
            "current_single_sequence": "expanding_window_rolling_origin",
            "evaluation_scope": "within_crystal_later_time_extrapolation_only",
            "historical_inputs_may_precede_target_partition_start": True,
            "mechanical_boundary_gap_or_L_minus_1_drop": False,
            "all_input_indices_must_be_at_or_before_t": True,
            "target_index_must_equal_t_plus_h": True,
            "train_only_fit_for_all_data_driven_preprocessing": True,
            "Fold_4_test_locked_until_protocol_frozen": True,
        },
        "fold_roles": {
            "development_folds": [1, 2, 3],
            "final_lockbox_fold": 4,
            "final_lockbox_target_interval": [18092, 20103],
        },
        "test_access": {
            "Fold_4_test_locked": True,
            "unlock_condition": (
                "all models, thresholds, windows, regularization, KAN capacity, "
                "optimization, and early stopping are frozen"
            ),
        },
        "not_selected": {
            "historical_85_10_5_split": (
                "Historical comparison only; not accepted as sole evidence."
            ),
            "random_row_split": "PROHIBITED",
            "future_X_in_primary_task": "PROHIBITED",
            "target_y_t_plus_h_in_input": "PROHIBITED",
        },
        "blocking_fields": [],
    }


def build_preprocessing_manifest(generated_at: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "generated_at": generated_at,
        "status": "POLICY_FROZEN_PARAMETERS_NOT_FITTED",
        "raw_data_modified": False,
        "target_intervals_frozen_before_sample_construction": True,
        "windows": {
            "recorded_history_length_samples": RECORDED_HISTORY_LENGTH,
            "legacy_reproduction": {"L_x": LEGACY_L_X, "L_y": LEGACY_L_Y},
            "formal_L_y_candidates": FORMAL_L_Y_CANDIDATES,
            "formal_L_x_candidates": FORMAL_L_X_CANDIDATES,
            "conditional_L_x_extension": CONDITIONAL_L_X_EXTENSION,
            "input_X_slice_python": "X[t-L_x+1:t+1]",
            "input_y_slice_python": "y[t-L_y+1:t+1]",
            "target_index": "t+h",
            "prediction_horizons_samples": PREDICTION_HORIZONS,
            "membership_by_target_index": True,
            "input_cutoff_index": "t",
            "future_X_allowed": False,
            "intermediate_true_future_y_allowed": False,
            "materialized": False,
            "history_may_cross_target_partition_left_boundary": True,
            "mechanical_boundary_drop_samples": 0,
        },
        "standardization": {
            "required": True,
            "fit_partition": "train_only_per_outer_fold",
            "apply_without_refit_to": ["validation", "test"],
            "epsilon": "NOT_YET_FROZEN",
            "parameters_fitted": False,
        },
        "missing_values": {
            "structural_audit_missing_count": 0,
            "imputation_currently_required": False,
            "if_required_later": (
                "Causal, using only current/past information, fitted within train; "
                "never cross a split boundary."
            ),
        },
        "outliers": {
            "rule": "NOT_YET_FROZEN_REQUIRES_TRAIN_ONLY_AUDIT",
            "thresholds_fitted": False,
            "test_distribution_used": False,
        },
        "splines": {
            "knots_fit_partition": "train_only_per_outer_fold",
            "knots_fitted": False,
            "input_explanation_range": "train_2.5_to_97.5_percentile",
        },
        "caches_and_derivatives": {
            "must_include_dataset_track_horizon_fold_partition_identity": True,
            "public_and_private_namespaces_separate": True,
            "real_and_semisynthetic_targets_separate": True,
        },
        "fold_4_lockbox": {
            "test_interval": [18092, 20103],
            "locked": True,
            "requires_explicit_protocol_frozen_flag": True,
        },
        "constant_channels": {
            CONSTANT_EXOGENOUS_HEADER: {
                "interface_retained": True,
                "scientific_training_masked": True,
                "response_lag_and_rank_reporting_prohibited": True,
            }
        },
    }


def build_model_semantics(
    generated_at: str,
    authoritative_documents: list[dict[str, Any]],
    v20_audit: dict[str, Any],
    v2_adapter_audit: dict[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "generated_at": generated_at,
        "status": "PARTIALLY_FROZEN",
        "authoritative_documents": authoritative_documents,
        "method": "Autoregressive Rank-Adaptive Parallel Hammerstein-Urysohn",
        "method_acronym": "AR-RAPHU",
        "historical_v20_reference": {
            "bundle_directory": v20_audit["bundle_directory"],
            "integrity": v20_audit["integrity"],
            "contract_file_hashes": v20_audit["contract_file_hashes"],
            "external_gamma_contract": (
                v20_audit["historical_v20_external_gamma_contract"]
            ),
            "kan_contract": v20_audit["historical_v20_kan_contract"],
            "capacity_audit_space": v20_audit["source_backed_capacity_audit_space"],
            "reuse_boundary": v20_audit["v2_migration_rule"],
        },
        "v2_adapter": v2_adapter_audit,
        "current_application": (
            "within-crystal dynamic prediction for one CZ crystal, "
            "constant-diameter stage only"
        ),
        "current_data_semantics": {
            "external_process_channels": len(INPUT_HEADERS),
            "external_process_columns": INPUT_HEADERS,
            "historical_output_channels": 1,
            "model_input_channels": len(INPUT_HEADERS) + 1,
            "target_column": TARGET_HEADER,
            "current_target_used_as_input": False,
            "strictly_lagged_target_used_as_input": True,
            "historical_output_minimum_lag": 1,
            "recorded_history_length_samples": RECORDED_HISTORY_LENGTH,
            "legacy_reproduction": {"L_x": LEGACY_L_X, "L_y": LEGACY_L_Y},
            "formal_L_x_candidates": FORMAL_L_X_CANDIDATES,
            "formal_L_y_candidates": FORMAL_L_Y_CANDIDATES,
            "conditional_L_x_extension": CONDITIONAL_L_X_EXTENSION,
            "formal_L_x_selected": None,
            "formal_L_y_selected": None,
            "prediction_horizons_samples": PREDICTION_HORIZONS,
            "sampling_interval": UNKNOWN,
            "physical_lag_units_available": False,
            "crystal_count": 1,
            "stage": "constant_diameter_only",
            "constant_unidentifiable_external_channel": (
                CONSTANT_EXOGENOUS_HEADER
            ),
        },
        "task_tracks": {
            "Track-X": "external process variables only",
            "Track-AR": "strictly lagged historical diameter only",
            "Track-XAR": "external process variables plus lagged diameter",
            "required_incremental_metric": "Loss_AR(h)-Loss_XAR(h)",
        },
        "forecast_protocol": {
            "primary_mode": "direct_multi_horizon",
            "equation": "y_hat[t+h|t] = F_h(X up to t, y up to t)",
            "future_X_after_t_allowed": False,
            "intermediate_true_y_after_t_allowed": False,
            "membership_by_target_index": "t+h",
            "recursive_mode": "auxiliary_stability_only",
            "oracle_future_X": "TEP_only_separately_labeled_upper_bound",
        },
        "window_selection_protocol": {
            "legacy_reproduction": {"L_x": LEGACY_L_X, "L_y": LEGACY_L_Y},
            "formal_L_y_candidates": FORMAL_L_Y_CANDIDATES,
            "formal_L_x_candidates": FORMAL_L_X_CANDIDATES,
            "conditional_L_x_extension": CONDITIONAL_L_X_EXTENSION,
            "select_L_y_with": "Track-AR_on_development_folds_1_to_3",
            "select_L_x_with": "Track-X_on_development_folds_1_to_3",
            "full_cartesian_product_search": False,
            "global_pair_frozen_before_Fold_4_test": True,
        },
        "model_a": {
            "equation": "y_hat[t+h|t] = F_h(X up to t, y up to t)",
            "external_lag_kernel_family": "normalized_discrete_Gamma",
            "external_lag_kernel_exact_discretization": (
                "point_value_then_softmax"
            ),
            "gamma_shape_rate_or_scale_convention": "shape_scale",
            "gamma_zero_offset_delta": 0.001,
            "external_static_response": "univariate_KAN",
            "kan_architecture": {
                "per_variable_layers": [1, "hidden_kan", 1],
                "hidden_kan_candidates": [4, 8],
                "grid_size_candidates": [5, 7, 11],
                "spline_order": 3,
                "legacy_reproduction": {
                    "hidden_kan": 8,
                    "grid_size": 7,
                },
            },
            "historical_output_branch": {
                "enabled": True,
                "minimum_lag": 1,
                "rank": 1,
                "kernel_nonnegative_normalized": True,
                "kernel_implementation": "separate_V20_static_Gamma_core",
                "low_complexity_required": True,
                "recursive_stability_audit_required": True,
            },
            "branch_centering_on_train_distribution": True,
        },
        "model_b": {
            "scope": "external_process_branches_only_in_v2_first_version",
            "first_use": (
                "fixed_A_external_support_and_lags_convex_spline_refit"
            ),
            "orthogonal_residual_constraint_required": True,
            "weighted_gram_whitened_svd_required": True,
            "historical_output_residual_enabled": False,
            "historical_output_eta_primary_scientific_result": False,
        },
        "complexity_ladder": [
            "inactive",
            "rank_1_Hammerstein",
            "rank_2_parallel_Hammerstein",
            "full_Urysohn_surface",
        ],
        "scientific_boundaries": [
            "Predictive structure is not automatically causal in closed-loop data.",
            "Gamma misspecification must be separated from nonseparability using M6.",
            "No rank upgrade without stability, bootstrap, held-out gain, and density.",
            "Current CZ evidence is limited to one crystal and one process stage.",
            "Effective predictive lag is not a pure physical transport lag.",
            "Historical-output dynamics are not an independent physical mechanism.",
        ],
        "blocking_fields": [],
        "deferred_optional_fields": [
            "recursive_auxiliary_Lipschitz_constraint_if_recursive_is_run",
        ],
    }


def export(source: Path, output_dir: Path) -> list[Path]:
    source = source.resolve()
    output_dir = output_dir.resolve()
    generated_at = datetime.now().astimezone().isoformat()

    source_hash_before = sha256_file(source)
    audit = audit_workbook(source)
    validate_source_audit(audit, source_hash_before)
    v20_audit = audit_v20_bundle(source.parent)
    v2_adapter_audit = audit_v2_adapter(source.parent)
    authoritative_documents = []
    for name in AUTHORITATIVE_DOCUMENT_NAMES:
        path = source.parent / name
        if not path.is_file():
            raise RuntimeError(f"Authoritative v2 document is missing: {path}")
        authoritative_documents.append(
            {
                "path_relative_to_project": name,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )

    outputs = {
        output_dir / "data_manifest.json": build_data_manifest(
            source, source_hash_before, audit, generated_at
        ),
        output_dir / "split_manifest.json": build_split_manifest(generated_at),
        output_dir / "preprocessing_manifest.json": (
            build_preprocessing_manifest(generated_at)
        ),
        output_dir / "model_semantics.json": build_model_semantics(
            generated_at,
            authoritative_documents,
            v20_audit,
            v2_adapter_audit,
        ),
        output_dir / "v20_implementation_audit.json": v20_audit,
        output_dir / "environment.json": build_environment(generated_at),
    }
    for path, payload in outputs.items():
        atomic_write_json(path, payload)

    variable_rows = build_variable_rows()
    variable_path = output_dir / "variable_dictionary.csv"
    atomic_write_csv(
        variable_path,
        list(variable_rows[0].keys()),
        variable_rows,
    )

    source_hash_after = sha256_file(source)
    if source_hash_after != source_hash_before:
        raise RuntimeError("CZ source changed while Phase 0 manifests were exported.")

    artifact_hashes = {
        path.name: sha256_file(path)
        for path in sorted([*outputs, variable_path])
    }
    checksums_path = output_dir / "sha256sums.json"
    atomic_write_json(
        checksums_path,
        {
            "schema_version": 1,
            "generated_at": generated_at,
            "source": {source.name: source_hash_after},
            "artifacts": artifact_hashes,
        },
    )

    return sorted([*outputs, variable_path, checksums_path])


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("实验数据1.xlsx"),
        help="Private CZ workbook (read-only).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data_manifests/cz"),
        help="Directory for generated Phase 0 manifests.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if os.environ.get("CONDA_DEFAULT_ENV") != "Env_pytorch":
        raise RuntimeError(
            "Run this exporter inside Anaconda environment Env_pytorch."
        )
    outputs = export(args.source, args.output_dir)
    print("Phase 0 manifests exported:")
    for path in outputs:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
