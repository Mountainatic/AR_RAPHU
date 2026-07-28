"""Frozen data semantics and access guards for CZ real-data protocol v1."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from hashlib import sha256
from math import floor
from pathlib import Path
from typing import Final

import numpy as np
from openpyxl import load_workbook


class CZProtocolError(RuntimeError):
    """Raised when a frozen CZ protocol invariant is violated."""


class FurnaceBLockedError(CZProtocolError):
    """Raised when the outer-furnace values are requested before R7."""


PROTOCOL_SCHEMA: Final = "CZ_REAL_DATA_PROTOCOL_V1"
FURNACE_A_SHA256: Final = (
    "c46e0d35d26903386fd80408f36660c4f8925a5dbc56c92527f020e433ef04de"
)
FURNACE_B_SHA256: Final = (
    "c3428966fe006572809156ee5e3f488264b8206b19b20887dcd00840bb26fbc3"
)
FURNACE_A_SAMPLES: Final = 20_103
FURNACE_B_SAMPLES: Final = 20_627

PRIMARY_INPUTS: Final = (
    "主加热功率",
    "晶升速度",
    "晶转速度",
    "埚升速度",
    "埚转速度",
)
TARGET: Final = "晶体直径"
SENSITIVITY_INPUT: Final = "加热元件温度"

EXPECTED_A_HEADERS: Final = (
    "加热元件温度",
    "主加热功率",
    "晶升速度",
    "晶转速度",
    "埚升速度",
    "埚转速度",
    "氩气流量设定",
    "晶体长度",
    "炉压",
    "晶体直径",
)
EXPECTED_B_HEADERS: Final = (
    "加热元件温度",
    "晶升速度",
    "晶转速度",
    "埚升速度",
    "埚转速度",
    "主加热功率",
    "晶体直径",
)

CANONICAL_BY_HEADER: Final = {
    "加热元件温度": "加热元件温度",
    "主加热功率": "主加热功率",
    "晶升速度": "晶升速度",
    "晶转速度": "晶转速度",
    "埚升速度": "埚升速度",
    "埚转速度": "埚转速度",
    "氩气流量设定": "氩气流量设定",
    "晶体长度": "晶体长度",
    "炉压": "炉压",
    "晶体直径": "晶体直径",
    "加热元件温度(°C)": "加热元件温度",
    "主加热功率(kW)": "主加热功率",
    "晶升速度(mm/min)": "晶升速度",
    "晶转速度(rpm)": "晶转速度",
    "埚升速度(mm/min)": "埚升速度",
    "埚转速度(rpm)": "埚转速度",
    "氩气流量设定(L/min)": "氩气流量设定",
    "晶体长度(mm)": "晶体长度",
    "炉压(Torr)": "炉压",
    "晶体直径(mm)": "晶体直径",
}

DIRECT_HORIZONS: Final = (1, 5, 15, 30, 60)
EXPLORATORY_HORIZONS: Final = (150,)
LX_GRID: Final = (32, 64, 128, 256, 512)
LY_GRID: Final = (8, 16, 32, 64, 128)
MTAU_GRID: Final = (16, 32, 48, 64)
MX_GRID: Final = (16, 20, 24, 28, 32)


@dataclass(frozen=True, slots=True)
class DevelopmentFold:
    fold: int
    nominal_train_start: int
    nominal_train_stop: int
    effective_train_stop: int
    validation_start: int
    validation_stop: int
    purge_gap: int

    def to_dict(self) -> dict[str, int]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class FurnaceData:
    inputs: np.ndarray
    target: np.ndarray
    temperature: np.ndarray
    canonical_headers: tuple[str, ...]
    source_sha256: str
    source_sheet: str


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def split_index(samples: int, fraction: float) -> int:
    """Map frozen split fractions using 0-based, left-closed floor indexing."""

    return floor(samples * fraction)


def purge_gap(*, L_x: int, L_y: int, h_max: int) -> int:
    if L_x < 1 or L_y < 1 or h_max < 1:
        raise CZProtocolError("L_x, L_y and h_max must be positive.")
    return max(L_x - 1, L_y - 1) + h_max


def build_development_folds(
    *,
    samples: int = FURNACE_A_SAMPLES,
    L_x: int,
    L_y: int,
    h_max: int = max(DIRECT_HORIZONS),
) -> tuple[DevelopmentFold, ...]:
    gap = purge_gap(L_x=L_x, L_y=L_y, h_max=h_max)
    boundaries = ((0.4, 0.5), (0.5, 0.6), (0.6, 0.7), (0.7, 0.8))
    folds: list[DevelopmentFold] = []
    for fold_id, (train_fraction, validation_fraction) in enumerate(
        boundaries, start=1
    ):
        nominal_stop = split_index(samples, train_fraction)
        effective_stop = nominal_stop - gap
        if effective_stop <= max(L_x, L_y):
            raise CZProtocolError("Purge leaves insufficient training history.")
        folds.append(
            DevelopmentFold(
                fold=fold_id,
                nominal_train_start=0,
                nominal_train_stop=nominal_stop,
                effective_train_stop=effective_stop,
                validation_start=nominal_stop,
                validation_stop=split_index(samples, validation_fraction),
                purge_gap=gap,
            )
        )
    return tuple(folds)


def confirmation_interval(samples: int = FURNACE_A_SAMPLES) -> tuple[int, int]:
    return split_index(samples, 0.8), samples


def assert_confirmation_access(*, protocol_frozen: bool, stage: str) -> None:
    if not protocol_frozen or stage not in {"R6", "R7", "R8", "R9"}:
        raise CZProtocolError(
            "Furnace-A internal confirmation is locked until frozen stage R6."
        )


def _read_numeric_sheet(
    path: Path,
    *,
    sheet_name: str,
    expected_headers: tuple[str, ...],
    expected_samples: int,
    expected_hash: str,
) -> tuple[np.ndarray, tuple[str, ...]]:
    actual_hash = file_sha256(path)
    if actual_hash != expected_hash:
        raise CZProtocolError(
            f"Workbook SHA256 mismatch for {path.name}: {actual_hash}"
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise CZProtocolError(f"Missing frozen sheet {sheet_name!r}.")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(next(rows))
        if headers != expected_headers:
            raise CZProtocolError(
                f"Header mismatch in {path.name}/{sheet_name}: {headers!r}"
            )
        values = np.asarray(list(rows), dtype=np.float64)
    finally:
        workbook.close()
    if values.shape != (expected_samples, len(expected_headers)):
        raise CZProtocolError(
            f"Shape mismatch in {path.name}/{sheet_name}: {values.shape}"
        )
    if not np.isfinite(values).all():
        bad = int(values.size - np.isfinite(values).sum())
        raise CZProtocolError(f"Non-finite cells found: {bad}.")
    canonical = tuple(CANONICAL_BY_HEADER[item] for item in headers)
    return values, canonical


def _project_primary(
    values: np.ndarray,
    canonical_headers: tuple[str, ...],
    *,
    source_hash: str,
    source_sheet: str,
) -> FurnaceData:
    by_name = {name: index for index, name in enumerate(canonical_headers)}
    return FurnaceData(
        inputs=np.column_stack([values[:, by_name[name]] for name in PRIMARY_INPUTS]),
        target=values[:, by_name[TARGET]].copy(),
        temperature=values[:, by_name[SENSITIVITY_INPUT]].copy(),
        canonical_headers=canonical_headers,
        source_sha256=source_hash,
        source_sheet=source_sheet,
    )


def load_furnace_a(path: str | Path) -> FurnaceData:
    source = Path(path)
    values, headers = _read_numeric_sheet(
        source,
        sheet_name="Sheet1",
        expected_headers=EXPECTED_A_HEADERS,
        expected_samples=FURNACE_A_SAMPLES,
        expected_hash=FURNACE_A_SHA256,
    )
    return _project_primary(
        values,
        headers,
        source_hash=FURNACE_A_SHA256,
        source_sheet="Sheet1",
    )


def load_furnace_b(
    path: str | Path,
    *,
    protocol_frozen: bool,
    stage: str,
) -> FurnaceData:
    if not protocol_frozen or stage not in {"R7", "R8", "R9"}:
        raise FurnaceBLockedError(
            "Furnace-B values are locked until the frozen R7 outer evaluation."
        )
    source = Path(path)
    values, headers = _read_numeric_sheet(
        source,
        sheet_name="Sheet2",
        expected_headers=EXPECTED_B_HEADERS,
        expected_samples=FURNACE_B_SAMPLES,
        expected_hash=FURNACE_B_SHA256,
    )
    return _project_primary(
        values,
        headers,
        source_hash=FURNACE_B_SHA256,
        source_sheet="Sheet2",
    )
