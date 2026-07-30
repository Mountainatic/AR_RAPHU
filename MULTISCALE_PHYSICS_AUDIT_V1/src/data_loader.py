"""Read-only two-rod workbook loader with frozen column semantics."""

from __future__ import annotations

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "主加热功率",
    "晶升速度",
    "埚升速度",
    "晶转速度",
    "埚转速度",
    "晶体直径",
)


@dataclass(frozen=True, slots=True)
class RodData:
    name: str
    columns: dict[str, np.ndarray]
    samples: int

    def get(self, name: str) -> np.ndarray:
        return self.columns[name]


@dataclass(frozen=True, slots=True)
class WorkbookData:
    path: Path
    sha256: str
    rods: dict[str, RodData]


def file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_sheet(
    workbook,
    sheet_name: str,
    *,
    required_columns: Iterable[str],
) -> RodData:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"MISSING_REQUIRED_SHEET:{sheet_name}")
    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    headers = tuple(str(value) if value is not None else "" for value in next(iterator))
    index = {name: position for position, name in enumerate(headers)}
    missing = [name for name in required_columns if name not in index]
    if missing:
        raise ValueError(f"MISSING_REQUIRED_COLUMNS:{sheet_name}:{missing}")
    values: dict[str, list[float]] = {name: [] for name in required_columns}
    for row_number, row in enumerate(iterator, start=2):
        for name in required_columns:
            value = row[index[name]]
            if value is None:
                raise ValueError(f"MISSING_VALUE:{sheet_name}:{row_number}:{name}")
            try:
                numeric = float(value)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"NON_NUMERIC_VALUE:{sheet_name}:{row_number}:{name}"
                ) from exc
            if not np.isfinite(numeric):
                raise ValueError(
                    f"NON_FINITE_VALUE:{sheet_name}:{row_number}:{name}"
                )
            values[name].append(numeric)
    arrays = {
        name: np.ascontiguousarray(column, dtype=np.float64)
        for name, column in values.items()
    }
    samples = len(next(iter(arrays.values())))
    if samples == 0 or any(len(column) != samples for column in arrays.values()):
        raise ValueError(f"INVALID_SHEET_LENGTH:{sheet_name}")
    return RodData(sheet_name, arrays, samples)


def load_workbook_data(
    path: str | Path,
    *,
    required_sheets: Iterable[str] = ("Sheet1", "Sheet2"),
    required_columns: Iterable[str] = REQUIRED_COLUMNS,
) -> WorkbookData:
    source = Path(path).resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        rods = {
            name: _read_sheet(
                workbook,
                name,
                required_columns=tuple(required_columns),
            )
            for name in required_sheets
        }
    finally:
        workbook.close()
    return WorkbookData(source, file_sha256(source), rods)
